"""
Importing excel of product details and doing some operations like
1. Creating set of all companies and no of products supplier_to_product
2. Creating set of total value of products (product quantity * price) for each company
3. Writing Total price of every row (product quantity * price) as 5th column in a new excel
"""

import openpyxl

inventory_excel = openpyxl.load_workbook("inventory.xlsx")
product_sheet = inventory_excel["Sheet1"]
supplier_to_product = {}
supplier_to_product_price = {}

for row in range(2,product_sheet.max_row + 1):
    supplier = product_sheet.cell(row, 4).value
    product_value = product_sheet.cell(row, 3).value
    product_quantity = product_sheet.cell(row, 2).value
    value_column = product_sheet.cell(row, 5)

    # Creating set of all companies and no of products
    if supplier in supplier_to_product:
        supplier_to_product[supplier] = supplier_to_product.get(supplier) + 1
    else:
        supplier_to_product[supplier] = 1

    # Creating set of total value of products (product quantity * price) for each company
    total_value = product_value * product_quantity
    if supplier in supplier_to_product_price:
        supplier_to_product_price[supplier] = supplier_to_product_price.get(supplier) + total_value
    else:
        supplier_to_product_price[supplier] = total_value

    # Saving total price value in the 5th column
    value_column.value = total_value

# Creating a new excel with all 5 columns
inventory_excel.save("inventory-updated.xlsx")
